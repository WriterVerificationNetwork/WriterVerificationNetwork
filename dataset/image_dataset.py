import glob
import os
from datetime import datetime
from pathlib import Path

import openpyxl
import torchvision
from PIL import Image, ImageOps
from torch.utils.data import Dataset
from tqdm import tqdm

from dataset.utils import resize_image, letters, MAX_WIDTH, MAX_HEIGHT, MAX_BIN_WIDTH, MAX_BIN_HEIGHT, letter_to_idx

tms_ignoring = {"60255", "60283", "60940", "61026", "61112", "61138"}


class ImageDataset(Dataset):

    def __init__(self, gt_dir, gt_binarized_dir, filter_neg_file, transforms, split_from, split_to):
        # Init folder dir
        self.gt_dir = gt_dir
        self.gt_binarized_dir = gt_binarized_dir
        self.transforms = transforms

        # Create image item
        temp_image_list = []
        for letter in letters:
            image_by_letter = glob.glob(os.path.join(self.gt_binarized_dir, f'{letter}_*.png'))
            sp_from, sp_to = int(len(image_by_letter) * split_from), int(len(image_by_letter) * split_to)
            temp_image_list.append(image_by_letter[sp_from: sp_to])

        negative_tm_list = create_negative_tm_pair(filter_neg_file, tms_ignoring)
        negative_tm_map = {f'{k}_{v}': (k, v) for k, v in negative_tm_list}
        self.image_list = []
        for image_by_letter in tqdm(temp_image_list):
            exising_keys = set({})
            for anchor in image_by_letter:
                positive_image_list = []
                negative_image_list = []
                anchor_tm = os.path.basename(anchor).split("_")[1]
                for img in image_by_letter:
                    img_tm = os.path.basename(img).split("_")[1]
                    if anchor == img or anchor_tm in tms_ignoring or img_tm in tms_ignoring:
                        continue
                    if anchor_tm == img_tm:
                        positive_image_list.append(img)
                    else:
                        if f'{anchor_tm}_{img_tm}' in negative_tm_map:
                            negative_image_list.append(img)
                for pos_image in positive_image_list:
                    for neg_image in negative_image_list:
                        anc_pos, pos_anc = anchor + pos_image + neg_image, pos_image + anchor + neg_image
                        if anc_pos not in exising_keys and pos_anc not in exising_keys:
                            self.image_list.append([anchor, pos_image, neg_image])
                            exising_keys.add(anc_pos)
                            exising_keys.add(pos_anc)

    def __getitem__(self, idx):
        # anchor
        anchor = os.path.basename(self.image_list[idx][0])
        img_anchor = get_image(os.path.join(self.gt_dir, anchor.split("_")[0], anchor),
                               self.transforms,
                               MAX_WIDTH,
                               MAX_HEIGHT,
                               False,
                               "img_anchor")
        bin_anchor = get_image(os.path.join(self.gt_binarized_dir, anchor),
                               self.transforms,
                               MAX_BIN_WIDTH,
                               MAX_BIN_HEIGHT,
                               True,
                               "bin_anchor")

        # positive image
        img = os.path.basename(self.image_list[idx][1])
        img_positive = get_image(os.path.join(self.gt_dir, img.split("_")[0], img),
                                 self.transforms,
                                 MAX_WIDTH,
                                 MAX_HEIGHT,
                                 False,
                                 "img_positive")
        bin_positive = get_image(os.path.join(self.gt_binarized_dir, img),
                                 self.transforms,
                                 MAX_BIN_WIDTH,
                                 MAX_BIN_HEIGHT,
                                 True,
                                 "bin_positive")

        # negative image
        img = os.path.basename(self.image_list[idx][2])
        img_negative = get_image(os.path.join(self.gt_dir, img.split("_")[0], img),
                                 self.transforms,
                                 MAX_WIDTH,
                                 MAX_HEIGHT,
                                 False,
                                 "img_negative")
        bin_negative = get_image(os.path.join(self.gt_binarized_dir, img),
                                 self.transforms,
                                 MAX_BIN_WIDTH,
                                 MAX_BIN_HEIGHT,
                                 True,
                                 "bin_negative")

        return {
            'symbol': letter_to_idx[anchor.split("_")[0]],
            'img_anchor': img_anchor,
            'bin_anchor': bin_anchor,
            'img_positive': img_positive,
            'bin_positive': bin_positive,
            'img_negative': img_negative,
            'bin_negative': bin_negative
        }

    def __len__(self) -> int:
        return len(self.image_list)


def get_image(image_path, data_transform, max_w, max_h, is_bin_img=False, name=""):
    with Image.open(image_path) as img:
        # Resize image to make sure image size is smaller than page size
        width, height = img.size
        ratio_w = max_w / width
        ratio_h = max_h / height
        scale = min(ratio_h, ratio_w)
        image = resize_image(img, scale).convert('RGB')
        image = ImageOps.invert(image)
        # Find the dominant color
        # dominant_color = bincount_app(np.asarray(img.convert("RGB")))
        # Add image to the background
        padding_image = Image.new(mode="RGB", size=(max_w, max_h), color=(0, 0, 0))
        padding_image.paste(image, box=(0, 0))
        # padding_image.save(name + ".png")

        if is_bin_img:
            padding_image = padding_image.convert("L")

    # Transform image
    img_tensor = data_transform(padding_image)

    return img_tensor


def create_negative_tm_pair(filter_file, tms_ignoring):
    filter_tm_file = Path(filter_file)
    filter_tm_object = openpyxl.load_workbook(filter_tm_file)
    sheet = filter_tm_object.active
    negative_tm_list = []
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=82, min_col=2, max_col=82, values_only=True)):
        for j, cell in enumerate(row):
            if cell == "X" and (str(2 + i) not in tms_ignoring) and (str(2 + j) not in tms_ignoring):
                negative_tm_list.append([str(sheet.cell(row=2 + i, column=1).value),
                                         str(sheet.cell(row=1, column=2 + j).value)])
                negative_tm_list.append([str(sheet.cell(row=1, column=2 + j).value),
                                         str(sheet.cell(row=2 + i, column=1).value)])
    return negative_tm_list


if __name__ == "__main__":
    ORIGINAL_FOLDER_DIR = "/home/mvu/Downloads/bt1_by_letters_20210824"
    GROUND_TRUTH_FOLDER_DIR = "/home/mvu/Downloads/bt1_by_letters_binarization"

    GOOD_GT_DATA_DIRNAME = GROUND_TRUTH_FOLDER_DIR + "/good"

    FILTER_TM_IMAGE_FILE_PATH = "/home/mvu/Downloads/Filter-Iliad-images.xlsx"
    start_time = datetime.now()
    dataset = ImageDataset(ORIGINAL_FOLDER_DIR,
                           GOOD_GT_DATA_DIRNAME,
                           FILTER_TM_IMAGE_FILE_PATH,
                           torchvision.transforms.Compose([torchvision.transforms.ToTensor()]),
                           split_from=0, split_to=0.8)
    print(len(dataset.image_list))
    end_time = datetime.now()
    print(end_time - start_time)
    print(dataset[0])
