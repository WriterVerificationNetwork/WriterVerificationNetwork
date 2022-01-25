import glob
import json
import os
import random
from datetime import datetime
from pathlib import Path

import openpyxl
import torchvision
from PIL import Image, ImageOps
from torch.utils.data import Dataset
from tqdm import tqdm

from dataset.utils import resize_image, letters, MAX_WIDTH, MAX_HEIGHT, MAX_BIN_WIDTH, MAX_BIN_HEIGHT, letter_to_idx


class TMDataset(Dataset):

    def __init__(self, gt_dir, gt_binarized_dir, filter_file, transforms, split_from, split_to,
                 unfold=False, min_n_sample_per_letter=0, min_n_sample_per_class=0, without_imgs=False):
        # Init folder dir
        self.gt_dir = gt_dir
        self.gt_binarized_dir = gt_binarized_dir
        self.transforms = transforms
        self.without_imgs = without_imgs
        self.anchor_tms = []

        # Create image item
        temp_image_list = []
        tm_map = {}
        letter_tm_map = {}
        total_imgs_removing_by_letter = 0
        for letter in letters:
            image_by_letter = glob.glob(os.path.join(self.gt_dir, f'{letter}_*.png'))
            if len(image_by_letter) < min_n_sample_per_letter:
                # Ignore if number of samples per letter is less than min_n_sample_per_letter
                total_imgs_removing_by_letter += len(image_by_letter)
                continue
            image_by_letter = sorted(image_by_letter)
            letter_tm_map[letter] = {}
            for img in image_by_letter:
                tm = os.path.basename(img).split("_")[1]
                if tm not in tm_map:
                    tm_map[tm] = []
                if tm not in letter_tm_map[letter]:
                    letter_tm_map[letter][tm] = []
                tm_map[tm].append(img)
                letter_tm_map[letter][tm].append(img)
            sp_from, sp_to = int(len(image_by_letter) * split_from), int(len(image_by_letter) * split_to)
            temp_image_list.append(image_by_letter[sp_from: sp_to])
        tm_filter = set([k for k, v in tm_map.items() if len(v) < min_n_sample_per_class])
        total_imgs_removing_by_class = sum([len(v) for k, v in tm_map.items() if len(v) < min_n_sample_per_class])
        print(f'Total number of images removed by letter lever filter: {total_imgs_removing_by_letter}')
        print(f'Total number of images removed by class lever filter: {total_imgs_removing_by_class}')

        positive_tms = {x: [x] for x in tm_map.keys()}
        negative_tms = {x: [] for x in tm_map.keys()}
        with open(filter_file) as f:
            triplet_filter = json.load(f)

        for item in triplet_filter['relations']:
            current_tm = item['category']
            for second_item in item['relations']:
                second_tm = second_item['category']
                relationship = second_item['relationship']
                if current_tm == '' or second_tm == '':
                    continue
                if relationship == 4:
                    negative_tms[current_tm].append(second_tm)
                    negative_tms[second_tm].append(current_tm)
                if relationship == 1:
                    positive_tms[current_tm].append(second_tm)
                    positive_tms[second_tm].append(current_tm)

        self.image_list = []
        for image_by_letter in tqdm(temp_image_list):
            for anchor in image_by_letter:
                anchor_tm = os.path.basename(anchor).split("_")[1]
                if anchor_tm in tm_filter:
                    continue
                anchor_letter = os.path.basename(anchor).split("_")[0]
                img_negative_tms = set(letter_tm_map[anchor_letter].keys()).intersection(negative_tms[anchor_tm])
                if len(img_negative_tms) > 0:
                    self.image_list.append(anchor)
                    self.anchor_tms.append(anchor_tm)
        self.letter_tm_map = letter_tm_map
        self.positive_tms = positive_tms
        self.negative_tms = negative_tms

    def __getitem__(self, idx):
        # anchor
        anchor_img = self.image_list[idx]
        anchor = os.path.basename(anchor_img)
        anchor_tm = anchor.split("_")[1]
        anchor_letter = anchor.split("_")[0]

        positive_tms = set(self.letter_tm_map[anchor_letter].keys()).intersection(self.positive_tms[anchor_tm])
        positive_tm = random.choice(list(positive_tms))
        positive_img = random.choice(self.letter_tm_map[anchor_letter][positive_tm])
        negative_tms = set(self.letter_tm_map[anchor_letter].keys()).intersection(self.negative_tms[anchor_tm])
        negative_tm = random.choice(list(negative_tms))
        negative_img = random.choice(self.letter_tm_map[anchor_letter][negative_tm])

        # anchor image
        moving_percent = random.randint(0, 10) / 10.
        anchor = os.path.basename(anchor_img)
        anchor_tm = anchor.split("_")[1]
        if self.without_imgs:
            return {
                'symbol': letter_to_idx[anchor.split("_")[0]],
                'anchor_path': anchor,
                'tm_anchor': anchor_tm
            }
        img_anchor = get_image(os.path.join(self.gt_dir, anchor),
                               self.transforms, MAX_WIDTH, MAX_HEIGHT, is_bin_img=False, mov=moving_percent)
        bin_anchor = get_image(os.path.join(self.gt_binarized_dir, anchor),
                               self.transforms, MAX_BIN_WIDTH, MAX_BIN_HEIGHT, is_bin_img=True, mov=moving_percent)

        # positive image
        moving_percent = random.randint(0, 10) / 10.
        img = os.path.basename(positive_img)
        img_positive = get_image(os.path.join(self.gt_dir, img),
                                 self.transforms, MAX_WIDTH, MAX_HEIGHT, is_bin_img=False, mov=moving_percent)
        bin_positive = get_image(os.path.join(self.gt_binarized_dir, img),
                                 self.transforms, MAX_BIN_WIDTH, MAX_BIN_HEIGHT, is_bin_img=True, mov=moving_percent)

        # negative image
        moving_percent = random.randint(0, 10) / 10.
        img = os.path.basename(negative_img)
        img_negative = get_image(os.path.join(self.gt_dir, img),
                                 self.transforms, MAX_WIDTH, MAX_HEIGHT, is_bin_img=False, mov=moving_percent)
        bin_negative = get_image(os.path.join(self.gt_binarized_dir, img),
                                 self.transforms, MAX_BIN_WIDTH, MAX_BIN_HEIGHT, is_bin_img=True, mov=moving_percent)

        return {
            'symbol': letter_to_idx[anchor.split("_")[0]],
            'img_anchor': img_anchor,
            'anchor_path': anchor,
            'tm_anchor': anchor_tm,
            'bin_anchor': bin_anchor,
            'img_positive': img_positive,
            'bin_positive': bin_positive,
            'img_negative': img_negative,
            'bin_negative': bin_negative
        }

    def __len__(self) -> int:
        return len(self.image_list)


def get_image(image_path, data_transform, max_w, max_h, is_bin_img=False, mov=0.):
    with Image.open(image_path) as img:
        # Resize image to make sure image size is smaller than page size
        width, height = img.size
        ratio_w = max_w / width
        ratio_h = max_h / height
        scale = min(ratio_h, ratio_w)
        image = resize_image(img, scale).convert('RGB')
        width, height = image.size
        if not is_bin_img:
            image = ImageOps.invert(image)
        # Find the dominant color
        # dominant_color = bincount_app(np.asarray(img.convert("RGB")))
        # Add image to the background
        padding_image = Image.new(mode="RGB", size=(max_w, max_h), color=(0, 0, 0))
        padding_image.paste(image, box=(int(mov * (max_w - width)), int(mov * (max_h - height))))
        # padding_image.save(name + ".png")

        if is_bin_img:
            padding_image = padding_image.convert("L")

    # Transform image
    if not is_bin_img:
        img_tensor = data_transform(padding_image)
    else:
        img_tensor = torchvision.transforms.ToTensor()(padding_image)

    return img_tensor


def create_negative_tm_pair(filter_file, tms_ignoring):
    filter_tm_file = Path(filter_file)
    filter_tm_object = openpyxl.load_workbook(filter_tm_file)
    sheet = filter_tm_object.active
    negative_tm_list = []
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=82, min_col=2, max_col=82, values_only=True)):
        for j, cell in enumerate(row):
            if cell == "X" and (str(2 + i) not in tms_ignoring) and (str(2 + j) not in tms_ignoring):
                negative_tm_list.append([str(sheet.cell(row=2 + i, column=1).value),
                                         str(sheet.cell(row=1, column=2 + j).value)])
                negative_tm_list.append([str(sheet.cell(row=1, column=2 + j).value),
                                         str(sheet.cell(row=2 + i, column=1).value)])
    return negative_tm_list


if __name__ == "__main__":
    ORIGINAL_FOLDER_DIR = "/home/mvu/Downloads/bt1_by_letters_20210824"
    GROUND_TRUTH_FOLDER_DIR = "/home/mvu/Downloads/bt1_by_letters_binarization"

    GOOD_GT_DATA_DIRNAME = GROUND_TRUTH_FOLDER_DIR + "/good"

    FILTER_TM_IMAGE_FILE_PATH = "/home/mvu/Downloads/Filter-Iliad-images.xlsx"
    start_time = datetime.now()
    dataset = TMDataset(ORIGINAL_FOLDER_DIR,
                           GOOD_GT_DATA_DIRNAME,
                           FILTER_TM_IMAGE_FILE_PATH,
                           torchvision.transforms.Compose([torchvision.transforms.ToTensor()]),
                           split_from=0, split_to=0.8)
    print(len(dataset.image_list))
    end_time = datetime.now()
    print(end_time - start_time)
    print(dataset[0])
